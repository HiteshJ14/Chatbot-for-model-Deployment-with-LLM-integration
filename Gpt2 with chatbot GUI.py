import os
import pickle
import tkinter as tk
from tkinter import ttk, scrolledtext
import random
import uuid
import joblib
import nltk
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize
from transformers import pipeline
import requests
import openpyxl
import pandas as pd
from datetime import datetime
from sklearn.preprocessing import LabelEncoder

# Initialize NLTK components
nltk.download('punkt')
stemmer = PorterStemmer()

# Define intents
intents = {
    "greetings": {
        "patterns": ["Hello", "Hi", "Hey", "How are you?", "Good morning", "Good evening"],
        "responses": ["Hello!", "Hi there!", "Hey!", "I'm good, thanks for asking!"]
    },
    "goodbye": {
        "patterns": ["Goodbye", "See you later", "Bye", "Take care"],
        "responses": ["Goodbye!", "See you later!", "Bye!", "Take care!"]
    },
    "thanks": {
        "patterns": ["Thanks", "Thank you", "I appreciate it"],
        "responses": ["You're welcome!", "No problem!", "Anytime!"]
    },
    "weather": {
        "patterns": ["What's the weather like?", "How's the weather?", "Tell me the weather", "Weather forecast"],
        "responses": []
    }
}

def tokenize_and_stem(sentence):
    tokens = word_tokenize(sentence)
    stems = [stemmer.stem(token.lower()) for token in tokens]
    return stems

def get_response(intents, user_input):
    tokenized_input = tokenize_and_stem(user_input)
    for intent, data in intents.items():
        for pattern in data['patterns']:
            tokenized_pattern = tokenize_and_stem(pattern)
            if tokenized_input == tokenized_pattern:
                return random.choice(data['responses'])
    return None

# Initialize text generation pipeline
generator = pipeline('text-generation', model='gpt2', tokenizer='gpt2', truncation=True, pad_token_id=50256)

def get_huggingface_response(user_input):
    inputs = generator.tokenizer(user_input, return_tensors='pt', truncation=True, padding='max_length', max_length=50)
    response = generator(model_inputs=inputs, max_length=50, num_return_sequences=1)
    reply = response[0]['generated_text'].strip()
    return reply

# OpenWeatherMap API configuration
OWM_API_KEY = "" # Enter your own OpenWeatherMAP API key
OWM_URL = "http://api.openweathermap.org/data/2.5/weather"

def get_weather(city):
    params = {
        'q': city,
        'appid': OWM_API_KEY,
        'units': 'metric'
    }
    response = requests.get(OWM_URL, params=params)
    weather_data = response.json()
    if response.status_code == 200:
        description = weather_data['weather'][0]['description']
        temperature = weather_data['main']['temp']
        return f"The weather in {city} is currently {description} with a temperature of {temperature}°C."
    else:
        return f"Sorry, I couldn't fetch the weather for {city}. Please try again."

# Questions for the bot
questions = [
    # (Add your questions here)
]

current_question = None
pending_fields = []
answers_data = []
answers = {}
attempts = 0
max_attempts = 3
chat_log = None  
model = None

def get_next_question(answers):
    for question in questions:
        if question["question"] not in answers:
            if "if_statements" in question and question["if_statements"]:
                for condition in question["if_statements"]:
                    if "if answer to" in condition:
                        condition_question = condition.split(" is ")[0].replace("if answer to ", "").strip()
                        condition_answer = condition.split(" is ")[1].replace("'", "").strip()
                        if answers.get(condition_question) != condition_answer:
                            break
                else:
                    return question
            else:
                return question
    return None

def load_model(file_path):
    with open(file_path, 'rb') as file:
        model = joblib.load(file)
    return model

def load_data_from_excel(excel_path):
    data = pd.read_excel(excel_path)
    return data

def encode_categorical_data(data):
    label_encoders = {}
    for column in data.select_dtypes(include=['object']).columns:
        label_encoders[column] = LabelEncoder()
        data[column] = label_encoders[column].fit_transform(data[column])
    return data, label_encoders

def prepare_data(data, model):
    data = data[model.feature_names_in_]
    data, label_encoders = encode_categorical_data(data)
    return data

def make_prediction(model, data):
    prediction = model.predict(data)
    return prediction

def handle_user_input(user_input):
    global current_question, pending_fields, attempts, chat_log, model

    if pending_fields:
        field = pending_fields.pop(0)
        answers_data.append((field["column"], user_input))
        answers[current_question["question"]] = user_input 

        if pending_fields:
            chat_log.insert(tk.END, f"Bot: Please provide the value for {pending_fields[0]['name']}.\n")
        else:
            current_question = get_next_question(answers)
            if current_question:
                if "fields" in current_question and current_question["fields"]:
                    pending_fields = current_question["fields"][:]
                    chat_log.insert(tk.END, f"Bot: Please provide the value for {pending_fields[0]['name']}.\n")
                elif current_question["choices"]:
                    question_text = current_question["question"] + " (" + ", ".join(current_question["choices"]) + ")"
                    chat_log.insert(tk.END, "Bot: " + question_text + "\n")
                else:
                    chat_log.insert(tk.END, "Bot: " + current_question["question"] + "\n")
            else:
                chat_log.insert(tk.END, "Bot: Thank you for answering the questions. Is there anything else I can help with?\n")
                save_to_excel()

                data = load_data_from_excel('answers.xlsx')
                data = prepare_data(data, model)
                prediction = make_prediction(model, data)
                chat_log.insert(tk.END, f"Bot: The model prediction is: {prediction[0]}\n")
        return

    if current_question:
        if current_question["question"] == "Are you registered?":
            if user_input.lower() == "yes":
                chat_log.insert(tk.END, "Bot: Please provide your Unique ID.\n")
                current_question = {"question": "Unique ID"}
                attempts = 0  
                return
            elif user_input.lower() == "no":
                chat_log.insert(tk.END, "Bot: Thank you for the information. You can proceed with other questions.\n")
                current_question = get_next_question(answers)
                if current_question:
                    if "fields" in current_question and current_question["fields"]:
                        pending_fields = current_question["fields"][:]
                        chat_log.insert(tk.END, f"Bot: Please provide the value for {pending_fields[0]['name']}.\n")
                    elif current_question["choices"]:
                        question_text = current_question["question"] + " (" + ", ".join(current_question["choices"]) + ")"
                        chat_log.insert(tk.END, "Bot: " + question_text + "\n")
                    else:
                        chat_log.insert(tk.END, "Bot: " + current_question["question"] + "\n")
                else:
                    chat_log.insert(tk.END, "Bot: Thank you for answering the questions. Is there anything else I can help with?\n")
                    save_to_excel()
                return
        elif current_question["question"] == "Unique ID":
            unique_id = user_input
            if check_unique_id_exists(unique_id):
                details = get_user_details(unique_id)
                chat_log.insert(tk.END, f"Bot: Your details: {details}\n")

                data = load_data_from_excel('answers.xlsx')
                data = prepare_data(data, model)
                prediction = make_prediction(model, data)
                chat_log.insert(tk.END, f"Bot: The model prediction is: {prediction[0]}\n")

                chat_log.insert(tk.END, "Bot: Is there anything else I can help with?\n")
                current_question = None
            else:
                attempts += 1
                if attempts < max_attempts:
                    chat_log.insert(tk.END, f"Bot: Unique ID not found. Please try again ({attempts}/{max_attempts} attempts).\n")
                else:
                    chat_log.insert(tk.END, "Bot: Maximum attempts reached. \nWould you like to register? (Yes/No)\n")
                    current_question = {"question": "Would you like to register?"}
            return
        elif current_question["question"] == "Would you like to register?":
            if user_input.lower() == 'yes':
                current_question = get_next_question(answers)
                if current_question:
                    if "fields" in current_question and current_question["fields"]:
                        pending_fields = current_question["fields"][:]
                        chat_log.insert(tk.END, f"Bot: Please provide the value for {pending_fields[0]['name']}.\n")
                    elif current_question["choices"]:
                        question_text = current_question["question"] + " (" + ", ".join(current_question["choices"]) + ")"
                        chat_log.insert(tk.END, "Bot: " + question_text + "\n")
                    else:
                        chat_log.insert(tk.END, "Bot: " + current_question["question"] + "\n")
                else:
                    chat_log.insert(tk.END, "Bot: Thank you for answering the questions. Is there anything else I can help with?\n")
                    save_to_excel()
                return
            elif user_input.lower() == 'no':
                chat_log.insert(tk.END, "Bot: Thank you. Have a great day!\n")
                current_question = None
            return

        if current_question["choices"]:
            if user_input in current_question["choices"]:
                answers[current_question["question"]] = user_input
                current_question = get_next_question(answers)
                if current_question:
                    if "fields" in current_question and current_question["fields"]:
                        pending_fields = current_question["fields"][:]
                        chat_log.insert(tk.END, f"Bot: Please provide the value for {pending_fields[0]['name']}.\n")
                    elif current_question["choices"]:
                        question_text = current_question["question"] + " (" + ", ".join(current_question["choices"]) + ")"
                        chat_log.insert(tk.END, "Bot: " + question_text + "\n")
                    else:
                        chat_log.insert(tk.END, "Bot: " + current_question["question"] + "\n")
                else:
                    chat_log.insert(tk.END, "Bot: Thank you for answering the questions. Is there anything else I can help with?\n")
                    save_to_excel()

                    data = load_data_from_excel('answers.xlsx')
                    data = prepare_data(data, model)
                    prediction = make_prediction(model, data)
                    chat_log.insert(tk.END, f"Bot: The model prediction is: {prediction[0]}\n")
            else:
                chat_log.insert(tk.END, "Bot: Please provide a valid option.\n")
        else:
            answers[current_question["question"]] = user_input
            current_question = get_next_question(answers)
            if current_question:
                if "fields" in current_question and current_question["fields"]:
                    pending_fields = current_question["fields"][:]
                    chat_log.insert(tk.END, f"Bot: Please provide the value for {pending_fields[0]['name']}.\n")
                elif current_question["choices"]:
                    question_text = current_question["question"] + " (" + ", ".join(current_question["choices"]) + ")"
                    chat_log.insert(tk.END, "Bot: " + question_text + "\n")
                else:
                    chat_log.insert(tk.END, "Bot: " + current_question["question"] + "\n")
            else:
                chat_log.insert(tk.END, "Bot: Thank you for answering the questions. Is there anything else I can help with?\n")
                save_to_excel()

                data = load_data_from_excel('answers.xlsx')
                data = prepare_data(data, model)
                prediction = make_prediction(model, data)
                chat_log.insert(tk.END, f"Bot: The model prediction is: {prediction[0]}\n")
    else:
        response = get_response(intents, user_input)
        if response:
            chat_log.insert(tk.END, "Bot: " + response + "\n")
        else:
            huggingface_response = get_huggingface_response(user_input)
            chat_log.insert(tk.END, "Bot: " + huggingface_response + "\n")

def save_to_excel():
    unique_id = str(uuid.uuid4())
    answers_data.append(("Unique ID", unique_id))
    answers_df = pd.DataFrame(answers_data, columns=["Question", "Answer"])
    answers_df.to_excel("answers.xlsx", index=False)

def check_unique_id_exists(unique_id):
    try:
        with open('answers.xlsx', 'rb') as file:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == unique_id:
                    return True
    except FileNotFoundError:
        return False
    return False

def get_user_details(unique_id):
    details = []
    try:
        with open('answers.xlsx', 'rb') as file:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == unique_id:
                    details = {headers[i]: row[i].value for i in range(len(headers))}
                    break
    except FileNotFoundError:
        return "Details not found"
    return details

# GUI Setup
root = tk.Tk()
root.title("Chatbot")

mainframe = ttk.Frame(root, padding="10")
mainframe.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

chat_log = scrolledtext.ScrolledText(mainframe, wrap=tk.WORD, state='disabled')
chat_log.grid(column=0, row=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))

user_input_entry = ttk.Entry(mainframe, width=50)
user_input_entry.grid(column=0, row=1, sticky=(tk.W, tk.E))
user_input_entry.focus()

def send_message(event=None):
    user_input = user_input_entry.get()
    chat_log['state'] = 'normal'
    chat_log.insert(tk.END, "You: " + user_input + "\n")
    chat_log['state'] = 'disabled'
    user_input_entry.delete(0, tk.END)
    handle_user_input(user_input)

send_button = ttk.Button(mainframe, text="Send", command=send_message)
send_button.grid(column=1, row=1, sticky=(tk.W, tk.E))

user_input_entry.bind("<Return>", send_message)

root.mainloop()
